import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def normalize_value(value) -> str:
    """Normalize values for change comparison: trim, collapse spaces, case-insensitive."""
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null", "nan"}:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def as_bool(value) -> bool:
    """Parse common truthy values."""
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "on"}


def safe_sheet_title(title: str) -> str:
    """Return an Excel-safe worksheet name (<=31 chars, no invalid chars)."""
    cleaned = re.sub(r"[:\\/?*\[\]]+", "_", title).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def safe_table_name(name: str) -> str:
    """Return a valid Excel table identifier."""
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    cleaned = cleaned.strip("_")
    if not cleaned:
        cleaned = "Table"
    if cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:250]


def load_csv_rows(input_path: Path) -> List[Dict[str, str]]:
    """Read CSV as list of dict rows."""
    with input_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def build_header_map(headers: Iterable[str]) -> Dict[str, str]:
    """Map normalized header names to actual names."""
    return {h.lower(): h for h in headers}


def pick_column(headers: Iterable[str], candidates: Iterable[str]) -> Optional[str]:
    """Return first existing header from candidate list (case-insensitive)."""
    header_map = build_header_map(headers)
    for candidate in candidates:
        actual = header_map.get(candidate.lower())
        if actual:
            return actual
    return None


def is_change(current_val, proposed_val) -> bool:
    """Detect dry-run change using normalized string comparison."""
    return normalize_value(current_val) != normalize_value(proposed_val)


def get_stable_key_column(headers: Iterable[str]) -> Optional[str]:
    """Find a stable row key, preferring employee identifiers."""
    return pick_column(
        headers,
        [
            "employeeID",
            "employeeId",
            "employee_id",
            "userPrincipalName",
            "distinguishedName",
            "id",
        ],
    )


def detect_current_proposed_pairs(headers: Iterable[str]) -> List[Tuple[str, str, str]]:
    """
    Detect generic current/proposed column pairs.
    Returns tuples: (logical_name, current_col, proposed_col).
    """
    header_map = build_header_map(headers)
    pairs = []
    for header in headers:
        if not header.lower().startswith("current"):
            continue
        suffix = header[len("current") :]
        if not suffix:
            continue
        for prefix in ("proposed", "new", "desired", "scheduled"):
            candidate = f"{prefix}{suffix}"
            actual = header_map.get(candidate.lower())
            if actual:
                logical = suffix.strip("_")
                pairs.append((logical, header, actual))
                break
    return pairs


def prepare_rows(
    source_rows: List[Dict[str, str]],
    columns: List[Tuple[str, str]],
    sort_key=None,
    reverse=False,
) -> List[List[str]]:
    """Project dict rows into ordered row lists for output."""
    rows = source_rows
    if sort_key is not None:
        rows = sorted(rows, key=sort_key, reverse=reverse)
    out = []
    for row in rows:
        out.append([row.get(src, "") for _, src in columns])
    return out


def write_table_sheet(
    wb: Workbook,
    title: str,
    headers: List[str],
    rows: List[List[str]],
    table_name: str,
    empty_note: Optional[str] = None,
) -> None:
    """Create worksheet, write rows, apply table, freeze header, and size columns."""
    ws = wb.create_sheet(title=safe_sheet_title(title))
    ws.freeze_panes = "A2"

    ws.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    data_rows = rows
    if not data_rows:
        note_row = [""] * len(headers)
        if empty_note and headers:
            note_row[0] = empty_note
        data_rows = [note_row]
    for row in data_rows:
        ws.append(row)

    end_row = ws.max_row
    end_col = ws.max_column
    ref = f"A1:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=safe_table_name(table_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    for col_idx in range(1, end_col + 1):
        letter = get_column_letter(col_idx)
        values = [ws.cell(row=r, column=col_idx).value for r in range(1, min(end_row, 1000) + 1)]
        max_len = max(len(str(v)) if v is not None else 0 for v in values)
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 70)


def build_readme_sheet(wb: Workbook) -> None:
    """Create README worksheet with usage/context notes."""
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Dry Run Change Report"
    ws["A1"].font = Font(size=14, bold=True)
    lines = [
        "This workbook is a dry run only. No directory changes were applied.",
        "",
        "Definitions:",
        "- Current: value currently in AD (or current source value in the dry-run input).",
        "- Proposed: value the sync process would write if changes were enabled.",
        "",
        "A user can appear in multiple tabs if multiple attributes would change.",
        "",
        "Risk flag definitions (Department_Changes):",
        "- currentEqualsManagerDept: TRUE when current department equals manager department.",
        "- proposedEqualsManagerDept: TRUE when proposed department equals manager department.",
        "- conflictsManagerDept: TRUE when currentEqualsManagerDept is TRUE and proposed differs from manager department.",
        "",
        "Missing/blank values are treated as empty for change detection.",
    ]
    row = 3
    for line in lines:
        ws[f"A{row}"] = line
        row += 1
    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A2"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Excel dry-run change report from CSV input.")
    parser.add_argument(
        "--input",
        default="adp_active_users_ad_current_vs_scheduled_department.csv",
        help="Input dry-run CSV path.",
    )
    parser.add_argument(
        "--output",
        default="dry_run_change_report.xlsx",
        help="Output Excel workbook path.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")

    rows = load_csv_rows(input_path)
    headers = list(rows[0].keys()) if rows else []
    key_col = get_stable_key_column(headers)

    wb = Workbook()
    build_readme_sheet(wb)

    # Column detection for named tabs
    current_employee_id = pick_column(headers, ["currentEmployeeID", "employeeIDCurrent"])
    proposed_employee_id = pick_column(headers, ["proposedEmployeeID", "employeeIDProposed", "newEmployeeID"])

    current_full_name = pick_column(headers, ["currentFullName", "fullNameCurrent"])
    proposed_full_name = pick_column(headers, ["proposedFullName", "fullNameProposed"])
    current_given_name = pick_column(headers, ["currentGivenName", "givenNameCurrent"])
    proposed_given_name = pick_column(headers, ["proposedGivenName", "givenNameProposed"])
    current_surname = pick_column(headers, ["currentSurname", "surnameCurrent", "currentSn"])
    proposed_surname = pick_column(headers, ["proposedSurname", "surnameProposed", "proposedSn"])

    current_title = pick_column(headers, ["currentTitle", "titleCurrent"])
    proposed_title = pick_column(headers, ["proposedTitle", "titleProposed"])

    current_dept = pick_column(headers, ["currentADDepartment", "currentDepartment", "departmentCurrent"])
    proposed_dept = pick_column(
        headers,
        [
            "proposedDepartmentV2",
            "proposedDepartmentFromScheduledUpdate",
            "proposedDepartment",
            "departmentProposed",
        ],
    )
    manager_name = pick_column(headers, ["userManager", "manager", "currentManager"])
    manager_dept = pick_column(headers, ["managerDepartment", "currentManagerDept"])
    missing_dept_signal = pick_column(headers, ["missingInADOrNoDept", "missingInAD", "missingOrNoDept"])

    current_manager = pick_column(headers, ["currentManager", "managerCurrent"])
    proposed_manager = pick_column(headers, ["proposedManager", "managerProposed", "newManager"])
    current_manager_dept = pick_column(headers, ["currentManagerDept"])
    proposed_manager_dept = pick_column(headers, ["proposedManagerDept"])

    # Build change sets
    employee_changes = []
    if current_employee_id and proposed_employee_id:
        for row in rows:
            if is_change(row.get(current_employee_id), row.get(proposed_employee_id)):
                employee_changes.append(row)

    name_changes = []
    if any([current_full_name and proposed_full_name, current_given_name and proposed_given_name, current_surname and proposed_surname]):
        for row in rows:
            changed = False
            if current_full_name and proposed_full_name:
                changed = changed or is_change(row.get(current_full_name), row.get(proposed_full_name))
            if current_given_name and proposed_given_name:
                changed = changed or is_change(row.get(current_given_name), row.get(proposed_given_name))
            if current_surname and proposed_surname:
                changed = changed or is_change(row.get(current_surname), row.get(proposed_surname))
            if changed:
                name_changes.append(row)

    title_changes = []
    if current_title and proposed_title:
        for row in rows:
            if is_change(row.get(current_title), row.get(proposed_title)):
                title_changes.append(row)

    department_changes = []
    if current_dept and proposed_dept:
        for row in rows:
            if is_change(row.get(current_dept), row.get(proposed_dept)):
                row["_currentEqualsManagerDept"] = (
                    bool(row.get(current_dept))
                    and bool(row.get(manager_dept or ""))
                    and normalize_value(row.get(current_dept)) == normalize_value(row.get(manager_dept or ""))
                )
                row["_proposedEqualsManagerDept"] = (
                    bool(row.get(proposed_dept))
                    and bool(row.get(manager_dept or ""))
                    and normalize_value(row.get(proposed_dept)) == normalize_value(row.get(manager_dept or ""))
                )
                row["_conflictsManagerDept"] = bool(row["_currentEqualsManagerDept"]) and not bool(row["_proposedEqualsManagerDept"])
                department_changes.append(row)
        department_changes.sort(
            key=lambda r: (
                not bool(r.get("_conflictsManagerDept")),
                not bool(r.get("_currentEqualsManagerDept")),
                str(r.get(key_col or "", "")),
            )
        )

    missing_or_no_dept = []
    if current_dept:
        for row in rows:
            current_blank = normalize_value(row.get(current_dept)) == ""
            missing_flag = as_bool(row.get(missing_dept_signal or ""))
            if current_blank or missing_flag:
                missing_or_no_dept.append(row)

    manager_changes = []
    if current_manager and proposed_manager:
        for row in rows:
            if is_change(row.get(current_manager), row.get(proposed_manager)):
                manager_changes.append(row)

    # Dept driver aggregation
    dept_driver_rows = []
    ref_field_col = pick_column(headers, ["departmentChangeReferenceField"])
    ref_value_col = pick_column(headers, ["departmentChangeReferenceValue"])
    if ref_field_col and ref_value_col:
        grouped = defaultdict(lambda: {"changeCount": 0, "conflictsManagerDeptCount": 0})
        for row in department_changes:
            key = (row.get(ref_field_col, ""), row.get(ref_value_col, ""))
            grouped[key]["changeCount"] += 1
            if row.get("_conflictsManagerDept"):
                grouped[key]["conflictsManagerDeptCount"] += 1
        for (field_val, ref_val), agg in grouped.items():
            count = agg["changeCount"]
            conflicts = agg["conflictsManagerDeptCount"]
            rate = (conflicts / count) if count else 0.0
            dept_driver_rows.append(
                {
                    "departmentChangeReferenceField": field_val,
                    "departmentChangeReferenceValue": ref_val,
                    "changeCount": count,
                    "conflictsManagerDeptCount": conflicts,
                    "conflictsManagerDeptRate": f"{rate:.2%}",
                }
            )
        dept_driver_rows.sort(key=lambda r: (-r["changeCount"], -int(r["conflictsManagerDeptCount"])))

    # Generic additional change tabs (non-address)
    explicit_cols = {
        col
        for col in [
            current_employee_id,
            proposed_employee_id,
            current_full_name,
            proposed_full_name,
            current_given_name,
            proposed_given_name,
            current_surname,
            proposed_surname,
            current_title,
            proposed_title,
            current_dept,
            proposed_dept,
            current_manager,
            proposed_manager,
        ]
        if col
    }
    address_keywords = {"address", "street", "city", "postal", "zip", "country", "state", "locality", "co", "st", "l"}
    generic_pairs = []
    for logical, cur_col, prop_col in detect_current_proposed_pairs(headers):
        if cur_col in explicit_cols or prop_col in explicit_cols:
            continue
        if any(k in logical.lower() for k in address_keywords):
            continue
        generic_pairs.append((logical, cur_col, prop_col))

    generic_change_tabs = []
    for logical, cur_col, prop_col in generic_pairs:
        changed_rows = [r for r in rows if is_change(r.get(cur_col), r.get(prop_col))]
        generic_change_tabs.append((logical, cur_col, prop_col, changed_rows))

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary.freeze_panes = "A2"
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A1"].font = Font(color="FFFFFF", bold=True)
    summary["B1"].font = Font(color="FFFFFF", bold=True)
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    summary["A1"].fill = fill
    summary["B1"].fill = fill
    total_rows = len(rows)
    unique_users = len({r.get(key_col, f"row_{i}") for i, r in enumerate(rows, start=1)}) if rows else 0

    summary_rows = [
        ("totalRows", total_rows),
        ("totalUsers", unique_users),
        ("EmployeeID_Changes", len(employee_changes)),
        ("Name_Changes", len(name_changes)),
        ("Title_Changes", len(title_changes)),
        ("Department_Changes", len(department_changes)),
        ("Manager_Changes", len(manager_changes)),
        ("Department_MissingOrNoDept", len(missing_or_no_dept)),
    ]
    if missing_dept_signal:
        summary_rows.append(("missingInADOrNoDeptFlagged", sum(1 for r in rows if as_bool(r.get(missing_dept_signal)))))
    if current_dept:
        summary_rows.append(("missingCurrentDepartment", sum(1 for r in rows if normalize_value(r.get(current_dept)) == "")))
    for logical, _, _, changed_rows in generic_change_tabs:
        summary_rows.append((f"{logical}_Changes", len(changed_rows)))

    for idx, (metric, value) in enumerate(summary_rows, start=2):
        summary.cell(row=idx, column=1, value=metric)
        summary.cell(row=idx, column=2, value=value)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 18

    # EmployeeID_Changes
    emp_cols = [
        ("employeeID", key_col or ""),
        ("fullName", pick_column(headers, ["fullName"]) or ""),
        ("currentEmployeeID", current_employee_id or ""),
        ("proposedEmployeeID", proposed_employee_id or ""),
        ("changeReason", pick_column(headers, ["employeeIDChangeReason", "changeReason"]) or ""),
        ("changeTrace", pick_column(headers, ["employeeIDChangeTrace", "changeTrace"]) or ""),
    ]
    emp_cols = [c for c in emp_cols if c[1]]
    write_table_sheet(
        wb,
        "EmployeeID_Changes",
        [c[0] for c in emp_cols] if emp_cols else ["note"],
        prepare_rows(employee_changes, emp_cols) if emp_cols else [],
        "EmployeeIDChangesTable",
        empty_note="No EmployeeID change rows or required columns missing.",
    )

    # Name_Changes
    name_cols = [
        ("employeeID", key_col or ""),
        ("currentFullName", current_full_name or ""),
        ("proposedFullName", proposed_full_name or ""),
        ("currentGivenName", current_given_name or ""),
        ("proposedGivenName", proposed_given_name or ""),
        ("currentSurname", current_surname or ""),
        ("proposedSurname", proposed_surname or ""),
        ("changeReason", pick_column(headers, ["nameChangeReason", "changeReason"]) or ""),
        ("changeTrace", pick_column(headers, ["nameChangeTrace", "changeTrace"]) or ""),
    ]
    name_cols = [c for c in name_cols if c[1]]
    write_table_sheet(
        wb,
        "Name_Changes",
        [c[0] for c in name_cols] if name_cols else ["note"],
        prepare_rows(name_changes, name_cols) if name_cols else [],
        "NameChangesTable",
        empty_note="No Name change rows or required columns missing.",
    )

    # Title_Changes
    title_cols = [
        ("employeeID", key_col or ""),
        ("fullName", pick_column(headers, ["fullName"]) or ""),
        ("currentTitle", current_title or ""),
        ("proposedTitle", proposed_title or ""),
        ("changeReason", pick_column(headers, ["titleChangeReason", "changeReason"]) or ""),
        ("changeTrace", pick_column(headers, ["titleChangeTrace", "changeTrace"]) or ""),
    ]
    title_cols = [c for c in title_cols if c[1]]
    write_table_sheet(
        wb,
        "Title_Changes",
        [c[0] for c in title_cols] if title_cols else ["note"],
        prepare_rows(title_changes, title_cols) if title_cols else [],
        "TitleChangesTable",
        empty_note="No Title change rows or required columns missing.",
    )

    # Department_Changes
    dept_columns = [
        ("employeeID", key_col or ""),
        ("fullName", pick_column(headers, ["fullName"]) or ""),
        ("title", pick_column(headers, ["title"]) or ""),
        ("currentADDepartment", current_dept or ""),
        ("proposedDepartmentFromScheduledUpdate", proposed_dept or ""),
        ("userManager", manager_name or ""),
        ("managerDepartment", manager_dept or ""),
        ("departmentChangeReferenceField", ref_field_col or ""),
        ("departmentChangeReferenceValue", ref_value_col or ""),
        ("departmentChangePrimaryReason", pick_column(headers, ["departmentChangePrimaryReason"]) or ""),
        ("departmentChangeReasonTrace", pick_column(headers, ["departmentChangeReasonTrace"]) or ""),
    ]
    dept_columns = [c for c in dept_columns if c[1]]
    dept_rows = prepare_rows(department_changes, dept_columns) if dept_columns else []
    for idx, row in enumerate(department_changes):
        if dept_columns:
            dept_rows[idx].extend(
                [
                    "TRUE" if row.get("_currentEqualsManagerDept") else "FALSE",
                    "TRUE" if row.get("_proposedEqualsManagerDept") else "FALSE",
                    "TRUE" if row.get("_conflictsManagerDept") else "FALSE",
                ]
            )
    dept_headers = [c[0] for c in dept_columns] if dept_columns else []
    dept_headers.extend(["currentEqualsManagerDept", "proposedEqualsManagerDept", "conflictsManagerDept"])
    write_table_sheet(
        wb,
        "Department_Changes",
        dept_headers if dept_headers else ["note"],
        dept_rows,
        "DepartmentChangesTable",
        empty_note="No Department change rows or required columns missing.",
    )

    # Department_MissingOrNoDept
    missing_columns = [
        ("employeeID", key_col or ""),
        ("fullName", pick_column(headers, ["fullName"]) or ""),
        ("title", pick_column(headers, ["title"]) or ""),
        ("currentADDepartment", current_dept or ""),
        ("proposedDepartmentFromScheduledUpdate", proposed_dept or ""),
        ("userManager", manager_name or ""),
        ("managerDepartment", manager_dept or ""),
        ("departmentChangeReferenceField", ref_field_col or ""),
        ("departmentChangeReferenceValue", ref_value_col or ""),
        ("departmentChangePrimaryReason", pick_column(headers, ["departmentChangePrimaryReason"]) or ""),
        ("departmentChangeReasonTrace", pick_column(headers, ["departmentChangeReasonTrace"]) or ""),
        ("missingInADOrNoDept", missing_dept_signal or ""),
    ]
    missing_columns = [c for c in missing_columns if c[1]]
    write_table_sheet(
        wb,
        "Department_MissingOrNoDept",
        [c[0] for c in missing_columns] if missing_columns else ["note"],
        prepare_rows(missing_or_no_dept, missing_columns) if missing_columns else [],
        "DepartmentMissingTable",
        empty_note="No missing/blank department rows or required columns missing.",
    )

    # Dept_ChangeDrivers
    driver_headers = [
        "departmentChangeReferenceField",
        "departmentChangeReferenceValue",
        "changeCount",
        "conflictsManagerDeptCount",
        "conflictsManagerDeptRate",
    ]
    driver_rows = [
        [
            r["departmentChangeReferenceField"],
            r["departmentChangeReferenceValue"],
            r["changeCount"],
            r["conflictsManagerDeptCount"],
            r["conflictsManagerDeptRate"],
        ]
        for r in dept_driver_rows
    ]
    write_table_sheet(
        wb,
        "Dept_ChangeDrivers",
        driver_headers,
        driver_rows,
        "DeptChangeDriversTable",
        empty_note="No department change driver data available.",
    )

    # Manager_Changes
    mgr_columns = [
        ("employeeID", key_col or ""),
        ("fullName", pick_column(headers, ["fullName"]) or ""),
        ("currentManager", current_manager or ""),
        ("proposedManager", proposed_manager or ""),
        ("currentManagerDept", current_manager_dept or manager_dept or ""),
        ("proposedManagerDept", proposed_manager_dept or ""),
        ("changeReason", pick_column(headers, ["managerChangeReason", "changeReason"]) or ""),
        ("changeTrace", pick_column(headers, ["managerChangeTrace", "changeTrace"]) or ""),
    ]
    mgr_columns = [c for c in mgr_columns if c[1]]
    write_table_sheet(
        wb,
        "Manager_Changes",
        [c[0] for c in mgr_columns] if mgr_columns else ["note"],
        prepare_rows(manager_changes, mgr_columns) if mgr_columns else [],
        "ManagerChangesTable",
        empty_note="No Manager change rows or required columns missing.",
    )

    # Generic additional tabs
    for logical, cur_col, prop_col, changed_rows in generic_change_tabs:
        tab_title = safe_sheet_title(f"{logical}_Changes")
        columns = [
            ("employeeID", key_col or ""),
            ("fullName", pick_column(headers, ["fullName"]) or ""),
            (f"current{logical}", cur_col),
            (f"proposed{logical}", prop_col),
        ]
        columns = [c for c in columns if c[1]]
        write_table_sheet(
            wb,
            tab_title,
            [c[0] for c in columns] if columns else ["note"],
            prepare_rows(changed_rows, columns) if columns else [],
            f"{logical}ChangesTable",
            empty_note=f"No {logical} change rows.",
        )

    output_path = Path(args.output)
    wb.save(output_path)
    print(f"Wrote workbook: {output_path}")
    print(f"Input rows: {len(rows)}")
    print("Change counts:")
    print(f"  EmployeeID: {len(employee_changes)}")
    print(f"  Name: {len(name_changes)}")
    print(f"  Title: {len(title_changes)}")
    print(f"  Department: {len(department_changes)}")
    print(f"  Manager: {len(manager_changes)}")
    print(f"  Missing/NoDept: {len(missing_or_no_dept)}")
    if generic_change_tabs:
        for logical, _, _, changed_rows in generic_change_tabs:
            print(f"  {logical}: {len(changed_rows)}")


if __name__ == "__main__":
    main()
